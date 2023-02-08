from bs4 import BeautifulSoup, NavigableString, Tag
import urllib.request
import os.path
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import time

def main():
    # Creating team class
    class Team:
        def __init__(self, name, win, score, short):
            self.name = name    # team name
            self.win = win      # bool if won
            self.score = score  # series score (will be 0, 1, or 2)
            self.short = short  # shortened name (for header)

    class Map:
        def __init__(self, name):
            self.name = name                # map name
        def setWinner(self, winner):
            self.winner = winner            # Team winner
        def setScore(self, totalScore):
            self.totalScore = totalScore    # array with final score
        def setHalfs(self, h1Score, h2Score):
            self.h1Score = h1Score          # array with half 1 score
            self.h2Score = h2Score          # array with half 2 score
        def setPick(self, isPick):
            self.isPick = isPick            # bool for map pick
        def setCT(self, isCT):
            self.isCT = isCT                # bool for ct first half
        def setOT(self, isOT, otScore):
            self.isOT = isOT                # bool for ot?
            self.otScore = otScore          # array for ot, will be None if no ot
        
    class Series:
        def __init__(self, totalScore, seriesWinner):
            self.totalScore = totalScore        # array of series score
            self.seriesWinner = seriesWinner    # series winner (string)
            
        def setMaps(self, maps):
            self.maps = maps                    # Map maps
        
        def totalT(self, tRounds):
            self.tRounds = tRounds              # list of total T rounds
        
        def totalCT(self, ctRounds):
            self.ctRounds = ctRounds            # list of total CT rounds
        
        roundDiff = 0

    wb = load_workbook('Spreadsheets/TEST.xlsx')

    # Fetching url to matchpage
    user_agent = 'Mozilla/5.0 (Windows; U; Windows NT 5.1; en-US; rv:1.9.0.7) Gecko/2009021910 Firefox/3.0.7'
    # url = input("HLTV Matchpage: ")
    url = "https://www.hltv.org/matches/2360116/villainous-vs-nouns-esl-challenger-league-season-43-north-america"
    headers={'User-Agent':user_agent,} 
    request=urllib.request.Request(url,None,headers)
    response = urllib.request.urlopen(request)
    data = response.read()
    
    matchpage = BeautifulSoup(data, 'lxml')

    # Getting team names
    team1 = matchpage.find('div', class_ ='team1-gradient').find('img', alt=True)

    team2 = matchpage.find('div', class_ ='team2-gradient').find('img', alt=True)

    sheets = wb.sheetnames
    sheetName = team1['alt'] + " vs. " + team2['alt']
    if sheetName not in sheets:
        wb.create_sheet(sheetName)
    ws = wb[sheetName]
    ws.delete_rows(1, 11)

    # Setting Team class
    # if team1 lost
    if (matchpage.find('div', class_ = 'team1-gradient').find(class_ = 'lost') != None):
        team1 = Team(team1['alt'], False, matchpage.find('div', class_ = 'team1-gradient').find(class_ = 'lost').text, team1['alt'][0:3].upper())
        team2 = Team(team2['alt'], True, matchpage.find('div', class_ = 'team2-gradient').find(class_ = 'won').text, team2['alt'][0:3].upper())
        series = Series([team1.score, team2.score], team2.name)
    # if team1 won
    else:
        team1 = Team(team1['alt'], True, matchpage.find('div', class_ = 'team1-gradient').find(class_ = 'won').text, team1['alt'][0:3].upper())
        team2 = Team(team2['alt'], False, matchpage.find('div', class_ = 'team2-gradient').find(class_ = 'lost').text, team2['alt'][0:3].upper())
        series = Series([team1.score, team2.score], team1.name)
    series.setMaps([])

    # Getting date
    date = matchpage.find('div', class_ = 'timeAndEvent').find('div', class_ = 'date').text
    dateArray = date.split()
    dateArray.remove('of')
    dateStr = team1.name + ' vs. ' + team2.name + ' ' + ' '.join(dateArray)
    ws.merge_cells("A1:E1")
    ws['A1'] = dateStr
    print(ws["A1"].value)

    # Getting vetoes
    getVetoes = matchpage.find_all('div', class_ = 'standard-box veto-box')
    getVetoes = getVetoes[-1]
    vetoes = getVetoes.find_all('div', class_ = None, text = True)
    final = []
    for v in vetoes:
        final.append(v.text.split())
    for f in final:
        f.pop(0)
        if len(f) == 3:
            if f[0] == team1.name:
                f[0] = team1.short
            else:
                f[0] = team2.short
            if f[1] == 'removed':
                f[1] = 'BAN'
            else:
                f[1] = 'PICK'
        else:
            f[2] = f[0]
            f[0] = "DECIDER"
            f[1] = ''

    # Pick team to find stats for
    # teamChoice = int(input("Which team's stats? [1] for " + team1.name + " [2] for " + team2.name + ": "))
    teamChoice = 2
    assert teamChoice == 1 or teamChoice == 2

    series.totalT = [0, 0]
    series.totalCT = [0, 0]

    # Map stats
    mapStats = matchpage.find('div', class_ = 'flexbox-column')
    mapCode = mapStats.find_all('div', class_ = 'mapholder')
    for map in mapCode:
        currentMap = Map(map.find('div', class_ = 'mapname').text)
        mapScore = map.find_all('div', class_ = 'results-team-score')
        mapScore = [mapScore[0].text.strip(), mapScore[1].text.strip()]
        if (teamChoice == 2):
            mapScore.reverse()
        currentMap.setScore(mapScore)
        # check if left side lost
        if (map.find('div', class_ = 'results-left lost') != None or map.find('div',  class_ = 'results-left lost pick') != None):
            currentMap.setWinner(team2.name)
        # left side wins
        else:
            currentMap.setWinner(team1.name)
        # check who picked map
        if (map.find('div', class_ = 'results-left lost pick') != None or map.find('div', class_ = 'results-left won pick') != None):
            if (teamChoice == 1):
                currentMap.setPick(True)
            else:
                currentMap.setPick(False)
        # If OT, parse through score differently
        if (len(map.find('div', class_ = 'results-center-half-score').find_all('span', class_ = True)) == 7):
            score = map.find('div', class_ = 'results-center-half-score').text.strip().replace(';', ')(').replace(' ', '')
            scoreByHalfs = score.split(')(')
            scoreByHalfs[0] = scoreByHalfs[0][1:]
            scoreByHalfs[2] = scoreByHalfs[2][:-1]
        else:
            score = map.find('div', class_ = 'results-center-half-score').text.strip().replace("(", '').replace(")", '').replace(' ', '')
            scoreByHalfs = score.split(';')

        # getting half scores
            
        half1 = scoreByHalfs[0].split(':')
        half1 = [int(i) for i in half1]
        half2 = scoreByHalfs[1].split(':')
        half2 = [int(i) for i in half2]
        
        if (teamChoice == 2):
            half1.reverse()
            half2.reverse()
        # checking if ot
        if (len(scoreByHalfs) == 3):
            ot = scoreByHalfs[2].split(':')
            ot = [int(i) for i in ot]
            if (teamChoice == 2):
                ot.reverse()
            currentMap.setOT(True, ot)
        else:
            currentMap.setOT(False, None)
        currentMap.setHalfs(half1, half2)

        # Check if first half was CT
        if teamChoice == 1 and '<span> (</span><span class="ct">' in str(map.find(class_ = 'results-center-half-score')) \
            or teamChoice == 2 and '<span> (</span><span class="ct">' not in str(map.find(class_ = 'results-center-half-score')):
            currentMap.setCT = True
            series.totalCT[0] += half1[0]
            series.totalCT[1] += half1[1]
            series.totalT[0] += half2[0]
            series.totalT[1] += half2[1]
        else:
            currentMap.setCT = False
            series.totalT[0] += half1[0]
            series.totalT[1] += half1[1]
            series.totalCT[0] += half2[0]
            series.totalCT[1] += half2[1]

        # Calculate Round Difference
        series.roundDiff += (int(mapScore[0]) - int(mapScore[1]))
        
        series.maps.append(currentMap)

    # Writing Pick/Ban Box to Excel
    ws.append([])
    ws.append(['PICK/BAN', 'MAP', 'SCORE', 'HALF 1', 'HALF 2', 'OT'])
    j = 0
    for i in range(len(final)):
        if final[i][1] == "BAN":
            mapBan = final[i][0] + ' ' + final[i][1]
            ws.append([mapBan, final[i][2]])
        elif final[i][1] == "PICK" or final[i][0] == "DECIDER":
            if final[i][1] == "PICK":
                mapPick = final[i][0] + ' ' + final[i][1]
            if final[i][0] == "DECIDER":
                mapPick = "DECIDER"
            mapScore = series.maps[j].totalScore[0] + '-' + series.maps[j].totalScore[1]
            half1 = str(series.maps[j].h1Score[0]) + '-' + str(series.maps[j].h1Score[1])
            half2 = str(series.maps[j].h2Score[0]) + '-' + str(series.maps[j].h2Score[1])
            if series.maps[j].setCT == True:
                half1 += " CT"
                half2 += " T"
            else:
                half1 += " T"
                half2 += " CT"
            toAppend = [mapPick, final[i][2], mapScore, half1, half2]
            if series.maps[j].isOT == True:
                toAppend.append(str(series.maps[j].otScore[0]) + '-' + str(series.maps[j].otScore[1]))
            ws.append(toAppend)
            j += 1
    score = series.totalScore[0] + "-" + series.totalScore[1]
    if series.totalScore[0] < series.totalScore[1]:
        score += " L"
    else:
        score += " W"
    ct = str(series.totalCT[0]) + "-" + str(series.totalCT[1]) + " CT"
    t = str(series.totalT[0]) + "-" + str(series.totalT[1]) + " T"
    ws.append(["OVERALL", score, "Diff: " + str(series.roundDiff), ct, t])
    ws.append([])
    
    # Getting detailed match stats link from matchpage & getting detailed match stats html
    getMatch = matchpage.find('div', class_ = 'small-padding stats-detailed-stats').find('a', href=True)
    detailedMatchStatsLink = 'https://www.hltv.org' + getMatch['href']
    print(detailedMatchStatsLink)
    request=urllib.request.Request(detailedMatchStatsLink, None, headers)
    response = urllib.request.urlopen(request)
    data = response.read()

    detailedStats = BeautifulSoup(data, 'lxml')

    mapStats = detailedStats.find('div', class_ = 'stats-match-maps')
    overall = detailedStats.find('div', class_ = 'match-info-box')

    allScoreboards = []
    tables = detailedStats.find_all('table', class_ = 'stats-table totalstats')
    # Get team 1 scoreboard
    thisMap = []
    curSB = []
    curRow = []
    for j in range(2):
        for row in tables[j].find_all('tr'):
            for col in row:
                if (col.text.strip() != ""):
                    curRow.append(col.text.strip())
            curSB.append(curRow.copy())
            curRow.clear()
        thisMap.append(curSB.copy())
        curSB.clear()
    allScoreboards.append(thisMap.copy())
    
    time.sleep(4.5)
    # Get link to economy page
    perfLink = detailedMatchStatsLink.split('/')
    perfLink.insert(5, "performance")
    perfLink = '/'.join(perfLink)
    print(perfLink)
    perfrequest=urllib.request.Request(perfLink, None, headers)
    perfresponse = urllib.request.urlopen(perfrequest)
    perfdata = perfresponse.read()
    perfStats = BeautifulSoup(perfdata, 'lxml')
    
    perfStats = perfStats.find('div', id = 'FIRST_KILL-content')
    perfStats = perfStats.find_all('td', class_ = True)
    fkMatrix = []
    fkRow = []
    count = 0
    for text in perfStats:
        if count == 6:
            count = 0
            fkMatrix.append(fkRow.copy())
            fkRow.clear()
        cell = text.text.replace(':', '-')
        cell = ' ' + cell
        fkRow.append(cell)
        count += 1
    fkMatrix.append(fkRow.copy())
    fkMatrix[0][0] = 'FK Matrix'

    # Getting urls for each map
    hrefCode = mapStats.find_all('a', class_ = 'col stats-match-map standard-box a-reset inactive', href=True)
    hrefs = []
    for i in hrefCode:
        hrefs.append('https://www.hltv.org' + i['href'])
    print(hrefs)

    # find stats for each map
    allClutches = []
    allPistols = []
    allSecond = []
    allForce = []
    allAnti = []
    allGun = []
    seriesEconomy = []
    for i in range(len(hrefs)):
        # print("PENISPENISPENISPENISPENISPENISPENISPENISPENISPENISPENISPENIS")
        time.sleep(4.5)
        request=urllib.request.Request(hrefs[i], None, headers)
        response = urllib.request.urlopen(request)
        data = response.read()
        statsPerMap = BeautifulSoup(data, 'lxml')

        # Get clutches
        # clutches = 'match-info-box-con', <div class="bold">Clutches won</div>
        getClutches = statsPerMap.find_all(class_ = "match-info-row")
        getClutches = getClutches[3].find('div', class_ =  'right').text
        getClutches = getClutches.split(":")
        for c in range(len(getClutches)):
            getClutches[c] = int(getClutches[c].strip())
        allClutches.append(getClutches)
        # get total clutches
        totalC = [0] * 2
        # print(getClutches)

        # Get pistol and second round stats
        # pistols + second = 'standard-box round-history-con' - check with title
        # class of round = 'round-history-outcome'
        getRoundHistory = statsPerMap.find_all('div', class_ = "round-history-team-row")
        if teamChoice == 1:
            getRoundHistory = getRoundHistory[0]
        else:
            getRoundHistory = getRoundHistory[1]
        getRoundHistory = getRoundHistory.find_all('img', class_ = 'round-history-outcome')
        pistol = []
        second = [0] * 4
        # Get Pistols
        if(getRoundHistory[0].attrs['title'] != ""):
            pistol.append(True)
        else:
            pistol.append(False)
        if(getRoundHistory[15].attrs['title'] != ""):
            pistol.append(True)
        else:
            pistol.append(False)

        # Get second round
        # convert stolen lose steal
        if (pistol[0] == True and getRoundHistory[1].attrs['title'] != ""):
            second[0] += 1
        else:
            second[1] += 1
        if (pistol[0] == False and getRoundHistory[1].attrs['title'] != ""):
            second[3] += 1
        else:
            second[2] += 1
        allSecond.append(second)

        # If on T first half, reverse order
        if (series.maps[i].setCT == False):
            pistol.reverse()
        totalP = [0] * 2
        allPistols.append(pistol)
        # print(pistols)

        # Get scoreboards

        # stats-table totalstats for tables
        tables = statsPerMap.find_all('table', class_ = 'stats-table totalstats')
        # Get team 1 scoreboard
        thisMap = []
        curSB = []
        curRow = []
        for j in range(2):
            for row in tables[j].find_all('tr'):
                for col in row:
                    if (col.text.strip() != ""):
                        curRow.append(col.text.strip())
                curSB.append(curRow.copy())
                curRow.clear()
            thisMap.append(curSB.copy())
            curSB.clear()
        allScoreboards.append(thisMap.copy())

        # force, anti, gun all in economy tab of matchpage
        
        # Get link to economy page
        econLink = hrefs[i].split('/')
        econLink.insert(5, "economy")
        econLink = '/'.join(econLink)

        newReq = urllib.request.Request(econLink, None, headers)
        newRes = urllib.request.urlopen(newReq)
        newDat = newRes.read()
        econPage = BeautifulSoup(newDat, 'lxml')

        thisMap.clear()
        curSB.clear()
        curRow.clear()
        getData = econPage.find_all('table', class_ = 'standard-box equipment-categories')
        # ct win, ct lose, t win, t lose
        ecos = [0] * 4
        guns = [0] * 4
        oneSideHalf = []
        oneHalf = []
        mapEconomy = []
        
        # # go through each cell and take type of round, then compare after, instead of direction comparison
        for k in range(2):
            for row in getData[k]: # gets to each row
                counter = 0
                for cell in row: # gets to each cell
                    if (str(cell).strip() != ""):
                        if counter > 1:
                            # ct side
                            if (str(cell).find('ct') != -1):
                                # check if eco/force
                                if (str(cell).find('Forcebuy') != -1 or \
                                    str(cell).find('Pistol') != -1):
                                    # check if won or lost
                                    if (str(cell).find('Win') != -1):
                                        oneSideHalf.append("ctForceW")
                                    else:
                                        oneSideHalf.append("ctForceL")
                                # last case = gun round
                                else:
                                    if (str(cell).find('Win') != -1):
                                        oneSideHalf.append("ctGunW")
                                    else:
                                        oneSideHalf.append("ctGunL")
                            # t side
                            else:

                                if (str(cell).find('Forcebuy') != -1 or \
                                    str(cell).find('Pistol') != -1):
                                    # check if won or lost
                                    if (str(cell).find('Win') != -1):
                                        oneSideHalf.append("tForceW")
                                    else:
                                        oneSideHalf.append("tForceL")
                                # last case = gun round
                                else:
                                    if (str(cell).find('Win') != -1):
                                        oneSideHalf.append("tGunW")
                                    else:
                                        oneSideHalf.append("tGunL")
                        counter += 1
                if (len(oneSideHalf) != 0):
                    oneHalf.append(oneSideHalf.copy())
                    oneSideHalf.clear()
            if (len(oneHalf) != 0):
                mapEconomy.append(oneHalf.copy())
                oneHalf.clear()
            if (len(mapEconomy) == 2):
                seriesEconomy.append(mapEconomy.copy())

    antis = []
    aC = 0
    # antieco stats, remember to remove from list after
    for currentMap in seriesEconomy: # currentMap = 1 map
        for sideHalf in currentMap:
            for teamHalf in sideHalf:
                if aC % 2 == 1:
                    antis.append(teamHalf.count('tForceL'))
                    antis.append(teamHalf.count('tForceW'))
                    antis.append(teamHalf.count('ctForceL'))
                    antis.append(teamHalf.count('ctForceW'))
                    
                    allAnti.append(antis.copy())
                    antis.clear()
                aC += 1
                
    newAnti = []
    tempor = [0] * 4
    for i in range(0, len(allAnti), 2):
        for j in range(4):
            tempor[j] = allAnti[i][j] + allAnti[i + 1][j]
        newAnti.append(tempor.copy())
    allAnti.clear()
    allAnti = newAnti.copy()

    # force and gun stats
    c = 0
    for currentMap in seriesEconomy: # currentMap = 1 map
        for sideHalf in currentMap:
            for teamHalf in sideHalf:
                if c % 2 == 0:
                    ecos[0] = teamHalf.count('ctForceW')
                    ecos[1] = teamHalf.count('ctForceL')
                    ecos[2] = teamHalf.count('tForceW')
                    ecos[3] = teamHalf.count('tForceL')
                    guns[0] = teamHalf.count('ctGunW')
                    guns[1] = teamHalf.count('ctGunL')
                    guns[2] = teamHalf.count('tGunW')
                    guns[3] = teamHalf.count('tGunL')
                    allForce.append(ecos.copy())
                    allGun.append(guns.copy())
                c += 1

    newForce = []
    tempo = [0] * 4
    for i in range(0, len(allForce), 2):
        for j in range(4):
            tempo[j] = allForce[i][j] + allForce[i + 1][j]
        newForce.append(tempo.copy())
    allForce.clear()
    allForce = newForce.copy()
    
    newGun = []
    tempo = [0] * 4
    for i in range(0, len(allGun), 2):
        for j in range(4):
            tempo[j] = allGun[i][j] + allGun[i + 1][j]
        newGun.append(tempo.copy())
    allGun.clear()
    allGun = newGun.copy()

    # reverse econ if teamchoice == 2
    if teamChoice == 2:
        # 0 <-> 3, 1 <-> 2
        # put all into 1 big list then iterate to switch
        switch = [allForce, allAnti, allGun]
        for i in range(len(switch)):
            for j in range(len(switch[i])):
                tmp = switch[i][j][0]
                switch[i][j][0] = switch[i][j][3]
                switch[i][j][3] = tmp
                tmp = switch[i][j][1]
                switch[i][j][1] = switch[i][j][2]
                switch[i][j][2] = tmp
    
    # calculate total clutches
    for cur in allClutches:
        for i in range(2):
            totalC[i] += cur[i]
    allClutches.append(totalC.copy())
    
    # calculate total pistols
    for cur in allPistols:
        for i in range(2):
            totalP[i] += int(cur[i])
    allPistols.append(totalP.copy())
    
    # calculate total forces
    totalF = [0] * 4
    for cur in allForce:
        for i in range(4):
            totalF[i] += cur[i]
    allForce.append(totalF.copy())
    
    # calculate total antieco
    totalA = [0] * 4
    for cur in allAnti:
        for i in range(4):
            totalA[i] += cur[i]
    allAnti.append(totalA.copy())
    
    # calculate total gun rounds
    totalG = [0] * 4
    for cur in allGun:
        for i in range(4):
            totalG[i] += cur[i]
    allGun.append(totalG.copy())
    
    # calculate total 2nd round
    total2 = [0] * 4
    for cur in allSecond:
        for i in range(4):
            total2[i] += cur[i]
    allSecond.append(total2.copy())
    
    # print('clutches')
    # print(allClutches)
    # print('pistols')
    # print(allPistols)
    # print('second')
    # print(allSecond)
    # # print('economy')
    # # print(seriesEconomy)
    # print('force')
    # print(allForce)
    # print('antieco')
    # print(allAnti)
    # print('gun')
    # print(allGun)
    # print('sb')
    # print(allScoreboards)
    # print('fk')
    # print(fkMatrix)
    
    # Save all to spreadsheet
        
    # Clutches
    ws.append(['Clutches:', 'Won:', 'Lost:', 'Total:'])
    for i in range(len(allClutches)):
        if i != len(allClutches) - 1:
            ws.append([str(series.maps[i].name), str(allClutches[i][0]), " " + str(allClutches[i][1]), str(allClutches[i][0]) + '-' + str(allClutches[i][1])])
        else:
            ws.append(['Total:', allClutches[i][0], allClutches[i][1], " " + str(allClutches[i][0]) + '-' + str(allClutches[i][1])])
    ws.append([])

    # Pistols
    ws.append(['Pistols:', 'Won:', 'Lost:', 'Total:'])
    for i in range(len(allPistols)):
        toAppend = []
        if i != len(allPistols) - 1: toAppend.append(series.maps[i].name)
        else: toAppend.append('Total:')
        toAppend.extend([int(allPistols[i][0]), int(allPistols[i][1]), " " + str(int(allPistols[i][0])) + '-' + str(int(allPistols[i][1]))])
        ws.append(toAppend)
    ws.append([])
    
    # Second
    ws.append(['2nd Round:', 'Convert:', 'Stolen:', 'Lose:', 'Steal:'])
    for i in range(len(allSecond)):
        toAppend = []
        if i != len(allSecond) - 1:
            ws.append([series.maps[i].name, allSecond[i][0], allSecond[i][1], allSecond[i][2], allSecond[i][3]])
        else:
            ws.append(['Total:', allSecond[i][0], allSecond[i][1], allSecond[i][2], allSecond[i][3]])
    ws.append([])
    

    # Eco/Force
    ws.append(['Eco/Force:', 'CT Won:', 'CT Lost:', 'T Won:', 'T Lost', 'Total:'])    
    for i in range(len(allForce)):
        toAppend = []
        if i != len(allForce) - 1: toAppend.append(series.maps[i].name)
        else: toAppend.append('Total:')
        toAppend.extend([allForce[i][0], allForce[i][1], allForce[i][2], allForce[i][3], " " + str(allForce[i][0] + allForce[i][2]) + '-' + str(allForce[i][1] + allForce[i][3])])
        ws.append(toAppend)
    ws.append([])
    
    # Anti Ecos
    ws.append(['Anti Ecos:', 'CT Won:', 'CT Lost:', 'T Won:', 'T Lost', 'Total:'])
    for i in range(len(allAnti)):
        toAppend = []
        if i != len(allAnti) - 1: toAppend.append(series.maps[i].name)
        else: toAppend.append('Total:')
        toAppend.extend([allAnti[i][0], allAnti[i][1], allAnti[i][2], allAnti[i][3], " " + str(allAnti[i][0] + allAnti[i][2]) + '-' + str(allAnti[i][1] + allAnti[i][3])])
        ws.append(toAppend)
    ws.append([])
    
    # # Gun Rounds
    ws.append(['Gun Rounds:', 'CT Won:', 'CT Lost:', 'T Won:', 'T Lost', 'Total:'])
    for i in range(len(allGun)):
        toAppend = []
        if i != len(allGun) - 1: toAppend.append(series.maps[i].name)
        else: toAppend.append('Total:')
        toAppend.extend([allGun[i][0], allGun[i][1], allGun[i][2], allGun[i][3], " " + str(allGun[i][0] + allGun[i][2]) + '-' + str(allGun[i][1] + allGun[i][3])])
        ws.append(toAppend)
    ws.append([])
    
    # FK Matrix
    for row in fkMatrix:
        ws.append(row)
    ws.append([])

    # Scoreboards
    scoreCount = -1
    for curMap in allScoreboards:
        if scoreCount == -1:
            ws.append(['TOTAL:'])
        else:
            ws.append([series.maps[scoreCount].name])
        for sb in curMap:
            for row in sb:
                ws.append(row)
        scoreCount += 1
        ws.append([])
    
    wb.save('Spreadsheets/TEST.xlsx')

if __name__ == "__main__":
    main()